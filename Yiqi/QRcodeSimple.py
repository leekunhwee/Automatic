#!/usr/bin/env python 
# -*- coding: utf-8 -*-

import os, sys
import xlrd
import openpyxl 
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import tkinter as tk
import tkinter.filedialog
import pyautogui
import pyperclip
import time
import json
import requests
import base64
import urllib.parse
import qrcode
from PIL import Image, ImageDraw, ImageFont  # pip install pillow
import numpy as np
import cv2

os.chdir(sys.path[0]) # 为了在 VScode 中调试时使用相对路径

# 原系统通过转码方式将姓名、身份证、电话转为二维码，所以这里模拟该过程生成二维码
def makeQRCode(userName, userCard, userTel):
	infoStr='{"username":"' + userName + '","usercard":"' + userCard +  '","usertel":"' + userTel +'","useraddr":""}'
	infoStr=base64.b64encode(infoStr.encode()) # base64加密
	infoStr=urllib.parse.quote(infoStr) # url加密
	qr = qrcode.QRCode(version=4, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=8, border=2)
	qr.add_data(infoStr)
	img = qr.make_image(fill_color="black", back_color="white")
	img_w, img_h = img.size
	# print(img_w, img_h)
	text ='姓名: ' + userName+'\n证件: '+userCard
	oriImg = Image.open('./source/background.png')

	oriImg.paste(img, (0, 13))
	draw = ImageDraw.Draw(oriImg)
	font_path = './source/msyh.ttc'

	font = ImageFont.truetype(font_path, 25) # 设置字体
	draw.text((img_w/20,img_h),text,fill='black',font=font)

	if not os.path.exists('./qrcodes'): os.mkdir('./qrcodes')
	oriImg.save('./qrcodes/'+ userName+'_'+userCard+".png")

# 定义选择 Excel 文件的程序
def findExcel():
	global source_file_name
	filename = tk.filedialog.askopenfilename(title='选择 Excel 文件', filetypes=[('Excel 表', '*.xls'),('Excel 表', '*.xlsx')]) # 限制文件选择类型	

	if filename != '':
		lb1.config(text = "您选择的Excel表是： "+filename);
		source_file_name = filename
		btn4.config(state = 'normal') # 激活截图选择按钮
	else:
		lb1.config(text = "您没有选择任何表格");

## 读取图像，解决imread不能读取中文路径的问题
def cv_imread(filePath):
    # 核心就是下面这句，一般直接用这句就行，直接把图片转为mat数据
    cv_img=cv2.imdecode(np.fromfile(filePath,dtype=np.uint8),-1)
    # imdecode读取的是rgb，如果后续需要opencv处理的话，需要转换成bgr，转换后图片颜色会变化
    # cv_img=cv2.cvtColor(cv_img,cv2.COLOR_RGB2BGR)
    return cv_img

def findShoot():
	global source_shoot_name
	filename = tk.filedialog.askopenfilename(title='选择 Png 截图', filetypes=[('屏幕截图', '*.png')]) # 限制文件选择类型

	if filename != '':
		lb2.config(text = "您选择的程序截图是："+filename);
		source_shoot_name = filename
		# source_shoot_name = cv_imread(filename)
		btn1.config(state = 'normal') # 激活确定按钮
	else:
		lb2.config(text = "您没有选择任何截图");

def fillName(name):
	pyautogui.click(x=left+width-width/8,y=top+barSize+blockSize/2,button='left')
	pyperclip.copy(name)
	pyautogui.hotkey('ctrl','v')

def fillSex(sex):
	if sex == '男':
		pyautogui.click(x=left+width-150,y=top+barSize+3*blockSize/2,button='left')
	else:
		pyautogui.click(x=left+width-80,y=top+barSize+3*blockSize/2,button='left')

def fillRelation(relation):
	pyautogui.click(x=left+width-width/8,y=top+barSize+5*blockSize/2,button='left')
	time.sleep(0.5)
	itemList=['本人','配偶','子女','父母','姐妹','兄弟','其他']
	for item in itemList:
		if(item!=relation):
			pyautogui.click(x=left+width/2,y=top+height-110-bottom,button='left')
		else:
			break
	pyautogui.click(x=left+width-30,y=top+height-360-bottom,button='left')

def fillLiveType(liveType):
	if liveType == '本地':
		pyautogui.click(x=left+width-195,y=top+barSize+7*blockSize/2,button='left')
	else:
		pyautogui.click(x=left+width-100,y=top+barSize+7*blockSize/2,button='left')

def fillLiveArea(liveArea):
	pyautogui.click(x=left+width-width/8,y=top+barSize+9*blockSize/2,button='left')
	time.sleep(0.5)
	pyautogui.click(x=left+width/2,y=top+height-450-bottom,button='left')
	
	#level0
	apiUrl='https://jc.ycqmhsjc.cn/checkApiInterface/app/getORGByPCode?'
	curArea="宜春市"
	lvl0Code='360900000000'
	lvl0Name=curArea

	#level1
	time.sleep(0.5)
	lvl1Res = requests.get(apiUrl+'code=' + lvl0Code + '&pCode=' + lvl0Code).text
	lvl1Data=json.loads(lvl1Res)['data']
	lvl1Code=''
	lvl1Name=''
	lvl1Len=len(lvl1Data)
	lvl1Idx=1

	for data in lvl1Data:
		areaName=data['name']
		areaCode=data['code']
		if(liveArea.startswith(curArea+areaName)):
			curArea=curArea+areaName
			lvl1Code=areaCode
			lvl1Name=areaName
			break
		lvl1Idx+=1

	if lvl1Idx<=9:
		pyautogui.click(x=left+width/2,y=top+height-450+55*(lvl1Idx-1)-bottom,button='left')
	elif lvl1Len-lvl1Idx<9:
		pyautogui.scroll(-2000)
		time.sleep(0.5)
		pyautogui.click(x=left+width/2,y=top+height-55*(lvl1Len-lvl1Idx)-bottom,button='left')
	else:
		pass

	#level2
	lvl2Res = requests.get(apiUrl+'code=' + lvl1Code + '&pCode=' + lvl1Code).text
	lvl2Data=json.loads(lvl2Res)['data']
	lvl2Code=''
	lvl2Name=''
	lvl2Len=len(lvl2Data)
	lvl2Idx=1
	for data in lvl2Data:
		areaName=data['name']
		areaCode=data['code']
		if(liveArea.startswith(curArea+areaName)):
			curArea=curArea+areaName
			lvl2Code=areaCode
			lvl2Name=areaName
			break
		lvl2Idx+=1

	if lvl2Idx<=9:
		pyautogui.click(x=left+width/2,y=top+height-450+55*(lvl2Idx-1)-bottom,button='left')
	elif lvl2Len-lvl2Idx<9:
		pyautogui.scroll(-2000)
		time.sleep(0.5)
		pyautogui.click(x=left+width/2,y=top+height-55*(lvl2Len-lvl2Idx)-bottom,button='left')
	else:
		pyautogui.scroll(-10)
		pyautogui.scroll(10)
		pyautogui.click(x=left+width-15,y=7*height/8,button='left')
		time.sleep(0.5)
		while lvl2Idx!=1:
			pyautogui.keyDown('down')
			lvl2Idx-=1
		pyautogui.click(x=left+width/2,y=top+height-450-bottom,button='left')

	#level3
	lvl3Res = requests.get(apiUrl+'code=' + lvl2Code + '&pCode=' + lvl2Code).text
	lvl3Data=json.loads(lvl3Res)['data']
	lvl3Code=''
	lvl3Name=''
	lvl3Len=len(lvl3Data)
	lvl3Idx=1
	for data in lvl3Data:
		areaName=data['name']
		areaCode=data['code']
		if(liveArea.startswith(curArea+areaName)):
			curArea=curArea+areaName
			lvl3Code=areaCode
			lvl3Name=areaName
			break
		lvl3Idx+=1
	# print(lvl3Idx)
	if lvl3Idx<=9:
		pyautogui.click(x=left+width/2,y=top+height-450+55*(lvl3Idx-1)-bottom,button='left')
	elif lvl3Len-lvl3Idx<9:
		pyautogui.scroll(-2000)
		time.sleep(0.5)
		pyautogui.click(x=left+width/2,y=top+height-55*(lvl3Len-lvl3Idx)-bottom,button='left')
	else:
		pyautogui.scroll(-10)
		pyautogui.scroll(10)
		pyautogui.click(x=left+width-15,y=7*height/8,button='left')
		time.sleep(0.5)
		while lvl3Idx!=1:
			pyautogui.keyDown('down')
			lvl3Idx-=1
		pyautogui.click(x=left+width/2,y=top+height-450-bottom,button='left')

def fillPhone(phone):
	pyautogui.click(x=left+width-width/8,y=top+barSize+11*blockSize/2,button='left')
	pyperclip.copy(phone)
	pyautogui.hotkey('ctrl','v')

def fillPeople(people):
	pyautogui.click(x=left+width-width/8,y=top+barSize+13*blockSize/2,button='left')
	time.sleep(0.5)
	itemList=['普通居民','学生','快递物流','医务人员','冷链人员','重点人群','省外返乡','住院病人','发热病人','陪护人员','红黄码人员','健康随访','公安','四类人群','集中隔离场所工作人员','入境人员','中风险地区人员','密接人员']
	for item in itemList:
		if(item!=people):
			pyautogui.click(x=left+width/2,y=top+height-110-bottom,button='left')
		else:
			break
	pyautogui.click(x=left+width-30,y=top+height-360-bottom,button='left')

def fillCardType(cardType):
	pyautogui.click(x=left+width-width/8,y=top+barSize+15*blockSize/2,button='left')
	time.sleep(0.5)
	itemList=['居民身份证','港澳台居民往地通行证','护照','台胞证']
	for item in itemList:
		if(item!=cardType):
			pyautogui.click(x=left+width/2,y=top+height-110-bottom,button='left')
		else:
			break
	pyautogui.click(x=left+width-30,y=top+height-360-bottom,button='left')

def fillCard(card):
	pyautogui.click(x=left+width-width/8,y=top+barSize+17*blockSize/2,button='left')
	pyperclip.copy(card)
	pyautogui.hotkey('ctrl','v')

def main(onerow):
	print(onerow)
	print('  ------------------------------------------')
	print("  若想中断循环，只需将鼠标移动至屏幕的某一角")
	print('  ------------------------------------------')
	pyautogui.click(x=left+width/2,y=top+height-150,button='left')
	time.sleep(0.5)
	fillName(onerow[0])
	time.sleep(0.5)
	fillSex(onerow[1])
	time.sleep(0.5)
	fillRelation(onerow[2])
	time.sleep(0.5)
	fillLiveType(onerow[3])
	time.sleep(0.5)
	fillLiveArea(onerow[4])
	time.sleep(0.5)
	fillPhone(onerow[5])
	time.sleep(0.5)
	fillPeople(onerow[6])
	time.sleep(0.5)
	fillCardType(onerow[7])
	time.sleep(0.5)
	fillCard(onerow[8])
	time.sleep(0.5)
	pyautogui.scroll(-1000)
	time.sleep(0.5)
	pyautogui.click(x=left+width/2,y=top+height-150,button='left')
	
	time.sleep(3)
	pyautogui.click(x=left+width/2,y=top+height-150,button='left')
	makeQRCode(onerow[0],onerow[8],onerow[5])

if __name__ == '__main__':
	print('\n')
	print('  -------------------------------------')
	print('* 欢迎使用“翼起防控”自动注册软件 V2.0 *')
	print('  -------------------------------------')
	print('\n')
	print('          ----------------------')
	print('        * 开发者：李健辉，王正仁 *')
	print('          ----------------------')
	print('  ---------------------------------------')
	print('* https://github.com/leekunhwee/Automatic *')
	print('  ---------------------------------------')
	print('\n')
	print('                 ----')
	print('               * 声明 *')
	print('                 ----')
	print('       软件仅限防疫人员录入信息\n       本软件不收集任何个人信息\n           不得用于商业用途\n           本人从未据此获利\n           严禁滥用这款软件')
	print('\n')
	os.system("pause")
	print('\n')
	print('  ------------------------------------')
	print('* 请提前按照要求准备受检人信息表和截图 *')
	print('  ------------------------------------')
	print('\n')
	os.system("pause")

	# 创建文件选择窗口
	window = tk.Tk()

	# 窗口命名 
	window.title('宜春加油!') 

	# 设定窗口大小(长 x 宽) 
	window.geometry('500x300')

	# 所选Excel表路径展示
	lb1 = tk.Label(window, text = '', wraplength = 500)
	lb1.place(x = 250, y = 50, anchor = 'center')

	# 所选screenshoot路径展示
	lb2 = tk.Label(window, text = '', wraplength = 500)
	lb2.place(x = 250, y = 80, anchor = 'center')

	# 完成文件选择按钮
	btn1 = tk.Button(window, text = "确定", bg = '#93FF93', command = window.destroy, state = 'disabled') # 选择文件之前的初始状态为未激活
	btn1.place(x = 220, y = 250, anchor = 'center')

	# 结束文件选择按钮
	btn2 = tk.Button(window, text = "取消", bg = '#FF9797', command = window.destroy) # 直接跳出
	btn2.place(x = 280, y = 250, anchor = 'center')

	# 选择文件按钮
	btn3 = tk.Button(window, text = "Excel表", bg = '#9393FF', command = findExcel)
	btn3.place(x = 200, y = 200, anchor = 'center')

	# 选择截图按钮
	btn4= tk.Button(window, text = "程序截图", bg = '#9393FF', command = findShoot, state = 'disabled') # 选择文件之前的初始状态为未激活
	btn4.place(x = 300, y = 200, anchor = 'center')

	# 获取Excel表格路径
	source_file_name = ''
	# 获取屏幕截图路径
	source_shoot_name = ''

	window.mainloop()
	
	if source_file_name == '': exit()
	if source_shoot_name == '': exit()

	# 截图尺寸初始化
	top=0
	left=0
	width=518 # 截图宽度像素
	height=975 # 截图高度像素
	blockSize=85 # 输入框高度
	barSize=55 # 菜单栏高度
	bottom=75 # 底部菜单栏高度
	# 获取截图的实际尺寸

	source_shoot_name = cv_imread(source_shoot_name)
	position=pyautogui.locateOnScreen(source_shoot_name)
	if(position==None):
		print('\n\n  -----------------------------------------------------')
		print("  请保证初始页面处于'添加成员'页面，且不被任何遮挡！！！")
		print('  -----------------------------------------------------\n\n')
		exit()
	(left,top,width,height)=(position.left,position.top,position.width,position.height)
	
	# 添加手动中断，将鼠标移动至屏幕某一角时，中断生效
	pyautogui.FAILSAFE = True

    # 判断打开文件的类型
	if(source_file_name[-1]=='s'): # 打开的是 .xls 文件
		sc = xlrd.open_workbook(filename = source_file_name)
        #通过索引获取表格sheet页
		sheet0 = sc.sheet_by_index(0)
		nrows=sheet0.nrows  #获取该表总行数
		ncols=sheet0.ncols  #获取该表总列数
		onerow = []
        # 获取首行
		for i in range(nrows):
			onerow = sheet0.row_values(i)
			for item in onerow:
				if item == '姓名':
					row_top = i
		onerow = []
		for i in range(row_top + 1, nrows):		# 直接拿出一行，本身就是 List
			onerow = sheet0.row_values(i) # 第 i+1 行
            # 数据清洗，把浮点型和整数型全部变为字符串
			for item in onerow:
				if type(item) == float:
					index_item = onerow.index(item)
					onerow[index_item] = int(onerow[index_item])
					onerow[index_item] = str(onerow[index_item])
				elif type(item) == int:
					index_item = onerow.index(item)
					onerow[index_item] = str(onerow[index_item])
			main(onerow)

	else:# 打开的是 .xlsx 文件
		sc = openpyxl.load_workbook(source_file_name)
		sheet0 = sc.worksheets[0]
		nrows = sheet0.max_row # 获得行数
		ncols = sheet0.max_column # 获得列数
		# 获取首行
		for row in sheet0.iter_rows(min_row = sheet0.min_row,    max_row = sheet0.max_row,
									min_col = sheet0.min_column, max_col = sheet0.max_column):
			for cell in row:
				if cell.value == '姓名':
					coord = coordinate_from_string(cell.coordinate) # returns ('A',1)
					col_top = column_index_from_string(coord[0]) # returns 1
					row_top = coord[1]
		# 将一行转化为 List
		for i in range(row_top, nrows):
			onerow = []
			for cell in list(sheet0.rows)[i]:  # 获取第i行的数据
				onerow.append(cell.value)
			# 数据清洗，把浮点型和整数型全部变为字符串
			for item in onerow:
				if type(item) == float:
					index_item = onerow.index(item)
					onerow[index_item] = int(onerow[index_item])
					onerow[index_item] = str(onerow[index_item])
				elif type(item) == int:
					index_item = onerow.index(item)
					onerow[index_item] = str(onerow[index_item])
			main(onerow)