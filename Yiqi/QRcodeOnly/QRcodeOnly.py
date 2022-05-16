#!/usr/bin/env python 
# -*- coding: utf-8 -*-

import xlrd
import openpyxl 
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import base64
import urllib.parse
import qrcode
import os
import tkinter as tk
import tkinter.filedialog

def makeQRCode(userName, userCard, userTel):
	infoStr='{"username":"' + userName + '","usercard":"' + userCard +  '","usertel":"' + userTel +'","useraddr":""}'
	infoStr=base64.b64encode(infoStr.encode())
	infoStr=urllib.parse.quote(infoStr)
	qr = qrcode.QRCode(version=4, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=8, border=2)
	qr.add_data(infoStr)
	img = qr.make_image(fill_color="black", back_color="white")
	if not os.path.exists('./qrcodes'): os.mkdir('./qrcodes')
	img.save('./qrcodes/'+ userName+'_'+userCard+".png")

# 定义选择 Excel 文件的程序
def findExcel():
	global source_file_name
	filename = tk.filedialog.askopenfilename(title='选择 Excel 文件', filetypes=[('Excel 表', '*.xls'),('Excel 表', '*.xlsx')]) # 限制文件选择类型

	if filename != '':
		lb1.config(text = "您选择的Excel表是："+filename);
		source_file_name = filename
		btn1.config(state = 'normal') # 激活确定按钮
	else:
		lb1.config(text = "您没有选择任何表格");

def main(onerow):
	print(onerow)
	makeQRCode(onerow[0],onerow[1],onerow[2])

if __name__ == '__main__':
	print('\n')
	print('  -----------------------------------------------')
	print('* 欢迎使用宜春“翼起防控”二维码自动生成软件 V1.0 *')
	print('  -----------------------------------------------')
	print('\n')
	print('              ----------------------')
	print('            * 开发者：李健辉，王正仁 *')
	print('              ----------------------')
	print('      ---------------------------------------')
	print('    * https://github.com/leekunhwee/Automatic *')
	print('      ---------------------------------------')
	print('\n')
	print('                       ----')
	print('                     * 声明 *')
	print('                       ----')
	print('             软件仅限防疫人员录入信息\n             本软件不收集任何个人信息\n                 不得用于商业用途\n                 本人从未据此获利\n                 严禁滥用这款软件')
	print('\n')
	os.system("pause")
	print('\n')
	print('          ------------------------------')
	print('        * 请提前按照要求准备受检人信息表 *')
	print('          ------------------------------')
	print('\n')
	os.system("pause")

	# 创建文件选择窗口
	window = tk.Tk()

	# 窗口命名 
	window.title('宜春加油!') 

	# 设定窗口大小(长 x 宽) 
	window.geometry('500x300')

	# 所选xlsx路径展示
	lb1 = tk.Label(window, text = '', wraplength = 500)
	lb1.place(x = 250, y = 50, anchor = 'center')

	# 所选screenshoot路径展示
	lb2 = tk.Label(window, text = '', wraplength = 500)
	lb2.place(x = 250, y = 80, anchor = 'center')

	# 完成文件选择按钮
	btn1 = tk.Button(window, text = "确定", bg = '#93FF93', command = window.destroy, state = 'disabled') # 选择文件之前的初始状态为未激活
	btn1.place(x = 220, y = 250, anchor = 'center')

	# 结束文件选择按钮
	btn2 = tk.Button(window, text = "取消", bg = '#FF9797', command = window.destroy) # 直接跳出
	btn2.place(x = 280, y = 250, anchor = 'center')

	# 选择文件按钮
	btn3 = tk.Button(window, text = "Excel表", bg = '#9393FF', command = findExcel)
	btn3.place(x = 250, y = 200, anchor = 'center')

	# 获取xlsx表格路径
	source_file_name = ''

	window.mainloop()
	
	if source_file_name == '': exit()
	
    # 判断打开文件的类型
	if(source_file_name[-1]=='s'): # 打开的是 .xls 文件
		sc = xlrd.open_workbook(filename = source_file_name)
        #通过索引获取表格sheet页
		sheet0 = sc.sheet_by_index(0)
		nrows=sheet0.nrows  #获取该表总行数
		ncols=sheet0.ncols  #获取该表总列数
		onerow = []
        # 获取首行
		for i in range(nrows):
			onerow = sheet0.row_values(i)
			for item in onerow:
				if item == '姓名':
					row_top = i
		onerow = []
		for i in range(row_top + 1, nrows):		# 直接拿出一行，本身就是 List
			onerow = sheet0.row_values(i) # 第 i+1 行
            # 数据清洗，把浮点型和整数型全部变为字符串
			for item in onerow:
				if type(item) == float:
					index_item = onerow.index(item)
					onerow[index_item] = int(onerow[index_item])
					onerow[index_item] = str(onerow[index_item])
				elif type(item) == int:
					index_item = onerow.index(item)
					onerow[index_item] = str(onerow[index_item])
			main(onerow)

	else:# 打开的是 .xlsx 文件
		sc = openpyxl.load_workbook(source_file_name)
		sheet0 = sc.worksheets[0]
		nrows = sheet0.max_row # 获得行数
		ncols = sheet0.max_column # 获得列数
		# 获取首行
		for row in sheet0.iter_rows(min_row = sheet0.min_row,    max_row = sheet0.max_row,
									min_col = sheet0.min_column, max_col = sheet0.max_column):
			for cell in row:
				if cell.value == '姓名':
					coord = coordinate_from_string(cell.coordinate) # returns ('A',1)
					col_top = column_index_from_string(coord[0]) # returns 1
					row_top = coord[1]
		# 将一行转化为 List
		for i in range(row_top, nrows):
			onerow = []
			for cell in list(sheet0.rows)[i]:  # 获取第i行的数据
				onerow.append(cell.value)
			# 数据清洗，把浮点型和整数型全部变为字符串
			for item in onerow:
				if type(item) == float:
					index_item = onerow.index(item)
					onerow[index_item] = int(onerow[index_item])
					onerow[index_item] = str(onerow[index_item])
				elif type(item) == int:
					index_item = onerow.index(item)
					onerow[index_item] = str(onerow[index_item])
			main(onerow)