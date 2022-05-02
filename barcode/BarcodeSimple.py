#!/usr/bin/env python 
# -*- coding: utf-8 -*-

import os 
import xlrd
import openpyxl 
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import tkinter as tk
import tkinter.filedialog
import pyautogui
import pyperclip
import time

#定义鼠标事件

#pyautogui库其他用法 https://blog.csdn.net/qingfengxd1/article/details/108270159

def mouseClick(clickTimes,lOrR,img,reTry):
    if reTry == 1:
        while True:
            location=pyautogui.locateCenterOnScreen(img,confidence=0.9)
            if location is not None:
                pyautogui.click(location.x,location.y,clicks=clickTimes,interval=0.2,duration=0.2,button=lOrR)
                break
            print("未找到匹配图片,0.1秒后重试，关闭程序按 Alt+F4 ")
            time.sleep(0.1)
    elif reTry == -1:
        while True:
            location=pyautogui.locateCenterOnScreen(img,confidence=0.9)
            if location is not None:
                pyautogui.click(location.x,location.y,clicks=clickTimes,interval=0.2,duration=0.2,button=lOrR)
            time.sleep(0.1)
    elif reTry > 1:
        i = 1
        while i < reTry + 1:
            location=pyautogui.locateCenterOnScreen(img,confidence=0.9)
            if location is not None:
                pyautogui.click(location.x,location.y,clicks=clickTimes,interval=0.2,duration=0.2,button=lOrR)
                print("重复")
                i += 1
            time.sleep(0.1)

#定义热键事件

#hotkey_get方法用来判断热键组合个数，并把热键传到对应的变量上newinput[0],[1],[2],[3]…只写了4个后续可以添加。
def hotkey_get(hk_g_inputValue):
            newinput = hk_g_inputValue.split(',')
            if len(newinput)==1: 
           			pyautogui.hotkey(hk_g_inputValue)
            elif len(newinput)==2:
           			pyautogui.hotkey(newinput[0],newinput[1])
            elif len(newinput)==3:
           			pyautogui.hotkey(newinput[0],newinput[1],newinput[2])
            elif len(newinput)==4:
           			pyautogui.hotkey(newinput[0],newinput[1],newinput[2],newinput[3])
                                                                   
#hotkey_Group方法调用hotkey_get方法，并判断其热键内容是否需要循环。
def hotkeyGroup(hotkey_reTry,hkg_inputValue):
    if hotkey_reTry == 1:
            hotkey_get(hkg_inputValue)                  
            print("执行了：",hkg_inputValue)
            time.sleep(0.1)
    elif hotkey_reTry == -1:
        while True:
            hotkey_get(hkg_inputValue)
            print("执行了：",hkg_inputValue)
            time.sleep(0.1)
    elif hotkey_reTry > 1:
        i = 1
        while i < hotkey_reTry + 1:
                hotkey_get(hkg_inputValue)
                print("执行了：",hkg_inputValue)
                i += 1
                time.sleep(0.1)
    
# 定义选择 xls 文件的程序
def find_Excel():
    global source_file_name
    # filename = tk.filedialog.askopenfilename(title='选择Excel文件', filetypes=[('Excel 表', '*.xls'),('Excel 表', '*.xlsx'), ('所有文件', '*')])
    filename = tk.filedialog.askopenfilename(title='选择 Excel 文件', filetypes=[('Excel 表', '*.xls'),('Excel 表', '*.xlsx')]) # 限制文件选择类型
    if filename != '':
        lb.config(text = "您选择的文件是："+filename);
        source_file_name = filename
        btn1.config(state = 'normal') # 激活确定按钮
    else:
        lb.config(text = "您没有选择任何文件");
        
# def exit_process():
    # exit()

#任务
def mainWork(sheet1,onerow):
    i = 1
    while i < sheet1.nrows:
        #取本行指令的操作类型
        cmdType = sheet1.row(i)[0]
        if cmdType.value == 1.0:
            #取图片名称
            img = sheet1.row(i)[1].value
            reTry = 1
            if sheet1.row(i)[2].ctype == 2 and sheet1.row(i)[2].value != 0:
                reTry = sheet1.row(i)[2].value
            mouseClick(1,"left",img,reTry)
            print("单击左键",img)
        #2代表双击左键
        elif cmdType.value == 2.0:
            #取图片名称
            img = sheet1.row(i)[1].value
            #取重试次数
            reTry = 1
            if sheet1.row(i)[2].ctype == 2 and sheet1.row(i)[2].value != 0:
                reTry = sheet1.row(i)[2].value
            mouseClick(2,"left",img,reTry)
            print("双击左键",img)
        #3代表右键
        elif cmdType.value == 3.0:
            #取图片名称
            img = sheet1.row(i)[1].value
            #取重试次数
            reTry = 1
            if sheet1.row(i)[2].ctype == 2 and sheet1.row(i)[2].value != 0:
                reTry = sheet1.row(i)[2].value
            mouseClick(1,"right",img,reTry)
            print("右键",img) 
        #4代表输入
        elif cmdType.value == 4.0:
            inputValue = sheet1.row(i)[1].value
            pyperclip.copy(inputValue)
            pyautogui.hotkey('ctrl','v')
            # time.sleep(0.5)
            print("输入:",inputValue)                                        
        #5代表等待
        elif cmdType.value == 5.0:
            #取图片名称
            waitTime = sheet1.row(i)[1].value
            time.sleep(waitTime)
            print("等待",waitTime,"秒")
        #6代表滚轮
        elif cmdType.value == 6.0:
            #取图片名称
            scroll = sheet1.row(i)[1].value
            pyautogui.scroll(int(scroll))
            print("滚轮滑动",int(scroll),"距离")     
       #7代表_热键组合
        elif cmdType.value == 7.0:
            #取重试次数,并循环。
            hotkey_reTry = 1
            if sheet1.row(i)[2].ctype == 2 and sheet1.row(i)[2].value != 0:
                hotkey_reTry = sheet1.row(i)[2].value
            inputValue = sheet1.row(i)[1].value
            hotkeyGroup(hotkey_reTry,inputValue)
            # time.sleep(0.5)
       #8代表_粘贴当前时间
        elif cmdType.value == 8.0:      
            #设置本机当前时间。
            localtime = time.strftime("%Y-%m-%d %H：%M：%S", time.localtime()) 
            pyperclip.copy(localtime)
            pyautogui.hotkey('ctrl','v')
            print("粘贴了本机时间:",localtime)
            # time.sleep(0.5)
       #9代表_系统命令集模式
        elif cmdType.value == 9.0:      
            wincmd = sheet1.row(i)[1].value
            os.system(wincmd)
            print("运行系统命令:",wincmd)
            # time.sleep(0.5) 
       #10代表输入名字
        elif cmdType.value == 10.0:
            inputValue = onerow[0]
            pyperclip.copy(inputValue)
            pyautogui.hotkey('ctrl','v')
            # time.sleep(0.5)
            print("输入:",inputValue) 
       #11代表输入身份证
        elif cmdType.value == 11.0:
            inputValue = onerow[1]
            pyperclip.copy(inputValue)
            pyautogui.hotkey('ctrl','v')
            # time.sleep(0.5)
            print("输入:",inputValue) 
       #12代表输入电话
        elif cmdType.value == 12.0:
            inputValue = onerow[2]
            pyperclip.copy(inputValue)
            pyautogui.hotkey('ctrl','v')
            # time.sleep(0.5)
            print("输入:",inputValue)   
       #13代表输入地址
        elif cmdType.value == 13.0:
            inputValue = onerow[3]
            pyperclip.copy(inputValue)
            pyautogui.hotkey('ctrl','v')
            # time.sleep(0.5)
            print("输入:",inputValue)
       #14代表输入图片文件名
        elif cmdType.value == 14.0:
            inputValue = onerow[0]+'_'+onerow[1]#用姓名和身份证命名条形码图片
            pyperclip.copy(inputValue)
            pyautogui.hotkey('ctrl','v')
            # time.sleep(0.5)
            print("输入:",inputValue)
        i += 1
        # input("Press <enter>")

if __name__ == '__main__':
    print('\n')
    print('       -------------------------')
    print('     * 欢迎使用自动注册软件 V1.1 *')
    print('       -------------------------')
    print('\n')
    print('            --------------')
    print('          * 开发者：李健辉 *')
    print('            --------------')
    print('       ------------------------')
    print('     * 部分代码参考 B 站 Up 主：*\n       * 不高兴就喝水、各种焱 *')
    print('         --------------------')
    print('     -----------------------------')
    print('   * https://github.com/leekunhwee *')
    print('     -----------------------------')
    print('\n')
    print('                 ----')
    print('               * 声明 *')
    print('                 ----')
    print('       软件仅限防疫人员录入信息\n       本软件不收集任何个人信息\n           不得用于商业用途\n           本人从未据此获利\n           严禁滥用这款软件')
    print('\n')
    os.system("pause")
    print('\n')
    print('         ----------------------------------')
    print('       * 请提前按照模板要求准备受检人信息表 *')
    print('         ----------------------------------')
    print('\n')
    os.system("pause")
    print('\n')
    print('     --------------------------------------------')
    print('   * 请在本电脑上通过桌面版微信打开条形码注册链接 *\n\n   * 停留在"添加受检人"页面，并保证该页面不被遮挡 *')
    print('     --------------------------------------------')
    print('\n')

    os.system("pause")
    print('\n')

    window = tk.Tk()

    # 窗口命名 
    window.title('上饶加油!') 

    # 设定窗口大小(长 x 宽) 
    window.geometry('500x300')

    source_file_name = ''

    # 所选文件路径展示
    lb = tk.Label(window, text = '', wraplength = 500)
    lb.place(x = 250, y = 50, anchor = 'center')
    # lb.pack()

    # 完成文件选择按钮
    btn1 = tk.Button(window, text = "确定", bg = '#93FF93', command = window.destroy, state = 'disabled') # 选择文件之前的初始状态为未激活
    btn1.place(x = 220, y = 250, anchor = 'center')

    # 结束文件选择按钮
    btn2 = tk.Button(window, text = "取消", bg = '#FF9797', command = window.destroy) # 直接跳出
    btn2.place(x = 280, y = 250, anchor = 'center')

    # 选择文件按钮
    btn3 = tk.Button(window, text = "选择Excel表", bg = '#9393FF', command = find_Excel)
    btn3.place(x = 250, y = 200, anchor = 'center')

    window.mainloop()
    
    if source_file_name == '':
        exit()
    
    print('\n')
    print('已选择文件：' + source_file_name)
    print('\n')

        # 判断打开文件的类型
    if(source_file_name[-1]=='s'): # 打开的是 .xls 文件
        sc = xlrd.open_workbook(filename = source_file_name)
        #通过索引获取表格sheet页
        sheet0 = sc.sheet_by_index(0)
        print(sheet0)
        nrows=sheet0.nrows  #获取该表总行数
        ncols=sheet0.ncols  #获取该表总列数
        
        onerow = []
        # 获取首行
        for i in range(nrows):
            onerow = sheet0.row_values(i)
            for item in onerow:
                if type(item) == float:
                    index_item = onerow.index(item)
                    onerow[index_item] = int(onerow[index_item])
                    onerow[index_item] = str(onerow[index_item])
                elif type(item) == int:
                    index_item = onerow.index(item)
                    onerow[index_item] = str(onerow[index_item])
            for item in onerow:
                if item == '姓名':
                    row_top = i
        
        onerow = []
        for i in range(row_top + 1, nrows):
            # 直接拿出一行，本身就是 List
            onerow = sheet0.row_values(i) # 第 i+1 行
            
            # 数据清洗，把浮点型和整数型全部变为字符串
            for item in onerow:
                if type(item) == float:
                    index_item = onerow.index(item)
                    onerow[index_item] = int(onerow[index_item])
                    onerow[index_item] = str(onerow[index_item])
                elif type(item) == int:
                    index_item = onerow.index(item)
                    onerow[index_item] = str(onerow[index_item])
            
            action_file = 'barcode.xls'
            #打开文件
            wb = xlrd.open_workbook(filename=action_file)
            #通过索引获取表格sheet页
            sheet1 = wb.sheet_by_index(0)
            
            mainWork(sheet1,onerow)


    else: # 打开的是 .xlsx 文件
        sc = openpyxl.load_workbook(source_file_name)
        sheet0 = sc.worksheets[0]
        nrows = sheet0.max_row # 获得行数
        ncols = sheet0.max_column # 获得列数
        
        # 获取首行
        for row in sheet0.iter_rows(min_row = sheet0.min_row,    max_row = sheet0.max_row,
                                    min_col = sheet0.min_column, max_col = sheet0.max_column):
            for cell in row:
                if cell.value == '姓名':
                    coord = coordinate_from_string(cell.coordinate) # returns ('A',1)
                    col_top = column_index_from_string(coord[0]) # returns 1
                    row_top = coord[1]
        
        # 将一行转化为 List
        for i in range(row_top, nrows):
            onerow = []
            for cell in list(sheet0.rows)[i]:  #获取第i行的数据
                onerow.append(cell.value)

            # 数据清洗，把浮点型和整数型全部变为字符串
            for item in onerow:
                if type(item) == float:
                    index_item = onerow.index(item)
                    onerow[index_item] = int(onerow[index_item])
                    onerow[index_item] = str(onerow[index_item])
                elif type(item) == int:
                    index_item = onerow.index(item)
                    onerow[index_item] = str(onerow[index_item])

            action_file = 'barcode.xls'
            #打开文件
            wb = xlrd.open_workbook(filename=action_file)
            #通过索引获取表格sheet页
            sheet1 = wb.sheet_by_index(0)
            
            mainWork(sheet1,onerow)